import pymongo
import pandas as pd
import numpy as np
import os 
import shutil
import xlrd
from xlutils.copy import copy
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

items  = ['A', 'B', 'C']
grades = ['1', '2', '3', '4', '5', '6', '7', '8']
groups = ['1', '2', '3']


file_name='110跳繩比賽v04_個人_成績登錄_0430_1355.xlsx'
export_file_name='110跳繩比賽v04_個人_成績登錄_0430_1355_未排序.xlsx'
myclient=pymongo.MongoClient('mongodb://192.168.88.12:27017/?readPreference=primary&appname=MongoDB%20Compass&directConnection=true&ssl=false')
mydb = myclient['djongo_test']
mycol=mydb['catalog_student']

def export_csv_nsort():
    if (os.path.isfile(file_name)):
        shutil.copyfile(file_name, export_file_name)        
    book = openpyxl.load_workbook(export_file_name,data_only = False)


    for item in items:
        for grade in grades:
            for group in groups:
                try:
                    sheet_name=item+grade+group
                    table=book[sheet_name]
                    
                    number_idx = 0
                    score_idx = 0
                    for i in range(1, table.max_column, 1):
                        if table.cell(row=1, column=i).value == '號碼':
                            number_idx = i
                        if table.cell(row=1, column=i).value == '成績':
                            score_idx = i

                    
                    for i in range(2, table.max_row, 1):
                        people = mycol.find({"std_No":table.cell(row=i, column=number_idx).value})
                        #table.cell(row=i, column=grade_idx, value=data[score])
                        for data in people:
                            table.cell(row=i, column=score_idx, value=data['score'])
                            #print(data['score'])
                            
                    
                    
                except Exception as e:
                   #print(item+grade+group, "ERROR!!")
                   #print(e)
                   continue
    book.save(export_file_name)
    
def data_export_sort():
    source_name = '110跳繩比賽v04_個人_成績登錄_0430_1355.xlsx'
    taget_name = '110跳繩比賽v04_個人_成績登錄_0430_1355_sorted.xlsx'
    
    shutil.copyfile(source_name, taget_name)
    
    book=load_workbook(taget_name, data_only=False)
    mapping = {6:1, 7:7, 8:6, 9:2, 10:8}
    for item in items:
        for grade in grades:
            for group in groups:
                try:
                    
                    #df = pd.read_excel(taget_name, sheet_name=item+grade+group)
                    data = mycol.find({"item":item, 'grade':grade, 'group':group}).sort("score", -1)
                    df = pd.DataFrame(list(data))
                    #print(df)
                    table = book[item+grade+group]
                    #print(type(table))
                    #print(table.values)
                    for i in range(0, df.shape[0]):
                        for j in range(6, 11):
                            table.cell(i+2, j).value = df.iloc[i,mapping[j]]
                            #print(df.iloc[i,:])
                        
                    
                    
                except Exception as e:
                   print(item+grade+group, "ERROR!!")
                   print(e)
                   continue
    book.save(taget_name)    
    
data_export_sort()
export_csv_nsort()